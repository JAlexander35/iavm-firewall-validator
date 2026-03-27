import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NVD_URL = "https://services.nvd.nist.gov/rest/json/cves/2.0"
CVE_PATTERN = re.compile(r"CVE-\d{4}-\d{4,7}", re.IGNORECASE)

COL_IAVM = "IAVM Notice"
COL_DESC = "Description"
COL_SEVERITY = "Severity"
COL_SCORE = "CVSS Base Score"
COL_FINDINGS = "Current Findings"
COL_CVES = "Related CVE(s)"
COL_NET = "Network Exploitable (Y/N)"
COL_PERF = "Performance Impact"
COL_SIG = "IPS Signature Available (Y/N)"
COL_MODE = "IPS Mode (Detect/Prevent/Inactive/N/A)"
COL_FW = "Firewall Coverage (Gateway/Cluster)"
COL_STATUS = "Mitigation Status"
COL_OWNER = "Owner"
COL_NOTES = "Notes / Actions"
COL_REC = "Recommendation"
COL_PRIORITY = "Analyst Priority"
COL_DETECT_REVIEW = "Recommend Detect Review"
COL_PREVENT_REVIEW = "Recommend Prevent Review"
COL_ACTION_TYPE = "Firewall Action Type"

CVE_BASE_COLUMNS = [
    COL_IAVM,
    COL_DESC,
    COL_SEVERITY,
    COL_SCORE,
    COL_CVES,
    COL_NET,
    COL_PERF,
    COL_SIG,
    COL_MODE,
    COL_FW,
    COL_STATUS,
    COL_OWNER,
    COL_NOTES,
    COL_REC,
]

IPS_METADATA_COLUMNS = {
    "follow up",
    "protection",
    "industry reference",
    "release date",
    "update date",
    "performance impact",
    "severity",
    "stig severity",
    "confidence level",
    "optimized",
    "default protection",
}

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
_SECTION_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")
_PRIMARY_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
_CRIT_FILL = PatternFill(fill_type="solid", fgColor="C00000")
_WARN_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
_GOOD_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
_BAD_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")


@dataclass
class CveResult:
    cve_id: str
    description: str = ""
    base_score: Optional[float] = None
    severity: str = ""
    attack_vector: str = ""
    network_exploitable: str = ""
    weakness_summary: str = ""
    cisa_kev: bool = False


def clean_str(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_header_name(value: str) -> str:
    s = clean_str(value)
    s = s.replace("\ufeff", "")
    s = s.replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip().lower()


COLUMN_ALIASES = {
    COL_IAVM: {"iavm notice", "iavm", "iavm block", "iavm notification"},
    COL_DESC: {"description", "details", "summary"},
    COL_SEVERITY: {"severity", "stig severity", "cat", "category"},
    COL_SCORE: {"cvss base score", "cvss", "cvss score", "base score"},
    COL_FINDINGS: {"current findings", "findings"},
    COL_CVES: {
        "related cve(s)",
        "related cves",
        "related cve",
        "cve(s)",
        "cves",
        "cve",
        "associated cve(s)",
        "associated cves",
    },
}


def canonicalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {}
    used_targets = set()

    for col in df.columns:
        norm = normalize_header_name(col)
        matched_target = None
        for target, aliases in COLUMN_ALIASES.items():
            if norm == normalize_header_name(target) or norm in aliases:
                matched_target = target
                break
        if matched_target and matched_target not in used_targets:
            rename_map[col] = matched_target
            used_targets.add(matched_target)
        else:
            rename_map[col] = clean_str(col)

    out = df.rename(columns=rename_map).copy()
    out.columns = [clean_str(c) for c in out.columns]
    out.attrs.update(df.attrs)
    return out


def split_cves(value) -> List[str]:
    if pd.isna(value):
        return []
    return [c.upper() for c in CVE_PATTERN.findall(str(value))]


def score_header_row(values: List[object]) -> int:
    score = 0
    normalized = [normalize_header_name(v) for v in values]
    for cell in normalized:
        if not cell:
            continue
        for target, aliases in COLUMN_ALIASES.items():
            if cell == normalize_header_name(target) or cell in aliases:
                score += 3
                break
        if "cve" in cell:
            score += 2
        if cell in {"description", "severity", "findings", "cvss base score", "iavm notice"}:
            score += 1
    return score


def read_excel_with_header_detection(path: str, max_scan_rows: int = 15) -> pd.DataFrame:
    preview = pd.read_excel(path, header=None, nrows=max_scan_rows)
    best_idx = 0
    best_score = -1

    for idx in range(len(preview)):
        row_values = preview.iloc[idx].tolist()
        score = score_header_row(row_values)
        if score > best_score:
            best_score = score
            best_idx = idx

    df = pd.read_excel(path, header=best_idx)
    df.attrs["detected_header_row"] = int(best_idx) + 1
    return df


def read_input_table(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        df = pd.read_csv(path)
        df.attrs["detected_header_row"] = 1
    else:
        df = read_excel_with_header_detection(path)
    return canonicalize_columns(df)


def fill_down_iavm_blocks(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    context_cols = [
        c
        for c in [COL_IAVM, COL_DESC, COL_SEVERITY, COL_SCORE, COL_FINDINGS, COL_STATUS, COL_OWNER]
        if c in out.columns
    ]
    if context_cols:
        out[context_cols] = out[context_cols].ffill()
    out.attrs.update(df.attrs)
    return out


def flatten_cve_blocks(df: pd.DataFrame) -> pd.DataFrame:
    df = canonicalize_columns(df)

    if COL_CVES not in df.columns:
        raise ValueError(
            "Could not find a CVE column. Expected something like 'Related CVE(s)'. "
            f"Columns found: {', '.join(map(str, df.columns))}"
        )

    df = fill_down_iavm_blocks(df)
    rows = []
    header_row = int(df.attrs.get("detected_header_row", 1))
    start_row = header_row + 1

    for original_row_number, (_, row) in enumerate(df.iterrows(), start=start_row):
        cves = split_cves(row.get(COL_CVES, ""))
        if not cves:
            continue
        for block_cve_index, cve in enumerate(cves, start=1):
            r = {col: row.get(col, "") for col in CVE_BASE_COLUMNS}
            r[COL_CVES] = cve
            r["Source Worksheet Row"] = original_row_number
            r["CVE Position In Source Cell"] = block_cve_index
            rows.append(r)

    if not rows:
        sample_values = [clean_str(v) for v in df[COL_CVES].head(10).tolist() if clean_str(v)]
        raise ValueError(
            "No CVE IDs were found in the CVE column after parsing. "
            f"Column used: '{COL_CVES}'. Sample values: {sample_values}"
        )

    out = pd.DataFrame(rows)
    for col in CVE_BASE_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    out.attrs.update(df.attrs)
    return out


def infer_profile_columns(df_ips: pd.DataFrame) -> List[str]:
    profile_cols = []

    for col in df_ips.columns:
        norm_col = normalize_header_name(col)
        if norm_col in IPS_METADATA_COLUMNS:
            continue

        values = (
            df_ips[col]
            .astype(str)
            .str.strip()
            .str.lower()
            .replace("nan", "", regex=False)
        )
        sample = [v for v in values.head(200).tolist() if clean_str(v)]
        if not sample:
            continue

        real_modes = {"prevent", "detect", "inactive", "n/a", "na"}
        boolish = {"true", "false", "yes", "no"}

        mode_hits = sum(v in real_modes for v in sample)
        bool_hits = sum(v in boolish for v in sample)

        if mode_hits > 0 and bool_hits == 0:
            profile_cols.append(col)

    return profile_cols


def normalize_mode(value: str) -> str:
    s = clean_str(value).lower()
    if not s:
        return ""
    if "prevent" in s:
        return "Prevent"
    if "detect" in s:
        return "Detect"
    if s in {"inactive", "n/a", "na", "disabled", "off", "false"}:
        return "Inactive"
    return clean_str(value)


def build_ips_lookup(df_ips: pd.DataFrame, profile_columns: List[str]) -> Dict[str, List[dict]]:
    if "Industry Reference" not in df_ips.columns:
        raise ValueError("IPS CSV/XLSX must contain an 'Industry Reference' column.")

    records = []
    for _, row in df_ips.iterrows():
        cves = split_cves(row.get("Industry Reference", ""))
        if not cves:
            continue

        profile_modes = {col: normalize_mode(row.get(col, "")) for col in profile_columns}

        for cve in cves:
            records.append(
                {
                    "cve_id": cve,
                    "protection": clean_str(row.get("Protection", "")),
                    "industry_reference": clean_str(row.get("Industry Reference", "")),
                    "performance": clean_str(row.get("Performance Impact", "")),
                    "severity": clean_str(row.get("Severity", "")),
                    "confidence": clean_str(row.get("Confidence Level", "")),
                    "optimized": normalize_mode(row.get("Optimized", "")),
                    "profiles": profile_modes,
                }
            )

    lookup: Dict[str, List[dict]] = {}
    for rec in records:
        lookup.setdefault(rec["cve_id"], []).append(rec)
    return lookup


def get_headers() -> Dict[str, str]:
    headers = {"User-Agent": "checkpoint-cve-iavm-enricher-v5/1.0"}
    api_key = os.getenv("NVD_API_KEY")
    if api_key:
        headers["apiKey"] = api_key
    return headers


def load_cache(path: Optional[str]) -> Dict[str, dict]:
    if not path or not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    return payload if isinstance(payload, dict) else {}


def save_cache(path: Optional[str], cache: Dict[str, dict]) -> None:
    if not path:
        return
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2, sort_keys=True)


def fetch_nvd_cve(session: requests.Session, cve_id: str, delay: float) -> CveResult:
    max_retries = 6

    for attempt in range(max_retries):
        try:
            response = session.get(
                NVD_URL,
                params={"cveId": cve_id},
                headers=get_headers(),
                timeout=30,
            )

            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After")
                if retry_after and str(retry_after).isdigit():
                    sleep_time = int(retry_after) + 1
                else:
                    sleep_time = min(30 * (2 ** attempt), 300)

                print(
                    f"WARN: 429 for {cve_id}; sleeping {sleep_time}s before retry",
                    file=sys.stderr,
                )
                time.sleep(sleep_time)
                continue

            response.raise_for_status()
            payload = response.json()
            vulns = payload.get("vulnerabilities", [])
            if not vulns:
                time.sleep(delay)
                return CveResult(cve_id=cve_id)

            cve = vulns[0].get("cve", {})
            description = ""
            for item in cve.get("descriptions", []):
                if item.get("lang") == "en":
                    description = item.get("value", "")
                    break

            weakness_summary = "; ".join(
                sorted(
                    {
                        d.get("value", "")
                        for weakness in cve.get("weaknesses", [])
                        for d in weakness.get("description", [])
                        if d.get("lang") == "en" and d.get("value")
                    }
                )
            )

            metrics = cve.get("metrics", {})
            base_score = None
            severity = ""
            attack_vector = ""

            for key in ("cvssMetricV40", "cvssMetricV31", "cvssMetricV30", "cvssMetricV2"):
                if key not in metrics or not metrics[key]:
                    continue
                metric = metrics[key][0]
                cvss = metric.get("cvssData", {})
                base_score = cvss.get("baseScore")
                severity = (
                    metric.get("baseSeverity")
                    or cvss.get("baseSeverity")
                    or metric.get("severity")
                    or ""
                )
                attack_vector = cvss.get("attackVector") or cvss.get("accessVector") or ""
                break

            cisa_kev = bool(cve.get("cisaExploitAdd"))
            network = "Y" if clean_str(attack_vector).upper() == "NETWORK" else "N"

            time.sleep(delay)
            return CveResult(
                cve_id=cve_id,
                description=description,
                base_score=base_score,
                severity=severity,
                attack_vector=attack_vector,
                network_exploitable=network,
                weakness_summary=weakness_summary,
                cisa_kev=cisa_kev,
            )

        except requests.RequestException as exc:
            if attempt == max_retries - 1:
                raise
            sleep_time = min(10 * (2 ** attempt), 120)
            print(
                f"WARN: request error for {cve_id}: {exc}; retrying in {sleep_time}s",
                file=sys.stderr,
            )
            time.sleep(sleep_time)

    raise RuntimeError(f"Failed to fetch {cve_id} after {max_retries} retries")


def summarize_signature_presence(
    matches: List[dict], profile_columns: List[str]
) -> Tuple[str, str, str, str, str]:
    if not matches:
        return "N", "N/A", "N/A", "", ""

    active_gateways = []
    all_modes = []
    perf_values = []
    protection_names = []

    for match in matches:
        protection_names.append(match.get("protection", ""))
        if match.get("performance"):
            perf_values.append(match["performance"])

        for col in profile_columns:
            raw_mode = match.get("profiles", {}).get(col, "")
            mode = normalize_mode(raw_mode)
            if mode:
                all_modes.append(mode)
            if mode in {"Detect", "Prevent"}:
                active_gateways.append(f"{col}:{mode}")

    if "Prevent" in all_modes:
        overall_mode = "Prevent"
    elif "Detect" in all_modes:
        overall_mode = "Detect"
    elif "Inactive" in all_modes:
        overall_mode = "Inactive"
    else:
        overall_mode = "N/A"

    coverage = ", ".join(sorted(set(active_gateways))) if active_gateways else "N/A"
    perf = ", ".join(sorted(set(v for v in perf_values if v)))
    protections = "; ".join(sorted(set(p for p in protection_names if p)))

    return "Y", overall_mode, coverage, perf, protections


def classify_priority(nvd: Optional[CveResult], sig_available: str, mode: str) -> str:
    severity = clean_str(nvd.severity if nvd else "").upper()
    network = clean_str(nvd.network_exploitable if nvd else "").upper()
    kev = bool(nvd.cisa_kev) if nvd else False

    if kev and network == "Y" and (sig_available == "N" or mode in {"Inactive", "N/A", "Detect"}):
        return "Critical"

    if network == "Y" and severity in {"CRITICAL", "HIGH"} and (
        sig_available == "N" or mode in {"Inactive", "N/A", "Detect"}
    ):
        return "High"

    if sig_available == "Y" and mode in {"Inactive", "Detect"}:
        return "Moderate"
    if network == "Y":
        return "Moderate"

    return "Low"


def derive_action_flags(sig_available: str, overall_mode: str, network: str, priority: str) -> Tuple[str, str]:
    sig = clean_str(sig_available).upper()
    mode = clean_str(overall_mode)
    net = clean_str(network).upper()
    prio = clean_str(priority)

    set_detect = ""
    set_prevent = ""

    # Broader Detect recommendation:
    # - signature exists but inactive / not active yet, or
    # - no matching signature in the provided export for a network-reachable CVE
    if net == "Y" and ((sig == "Y" and mode in {"Inactive", "N/A", ""}) or sig == "N"):
        set_detect = "✓"

    # Strict Prevent recommendation:
    # - matched signature already in Detect
    # - network reachable
    # - priority High/Critical
    if sig == "Y" and net == "Y" and mode == "Detect" and prio in {"Critical", "High"}:
        set_prevent = "✓"

    return set_detect, set_prevent


def choose_recommendation(
    row: pd.Series,
    nvd: Optional[CveResult],
    sig_available: str,
    overall_mode: str,
    perf: str,
) -> Tuple[str, str, str]:
    desc = clean_str((nvd.description if nvd else "") or row.get(COL_DESC, ""))
    network = clean_str((nvd.network_exploitable if nvd else "") or row.get(COL_NET, ""))
    severity = clean_str((nvd.severity if nvd else "") or row.get(COL_SEVERITY, ""))
    action_bits = []

    if network == "Y":
        action_bits.append("Network-reachable issue based on CVSS/NVD attack vector")
    elif network == "N":
        action_bits.append("No network attack vector identified from CVSS/NVD")
    else:
        action_bits.append("Network exploitability requires analyst review")

    if nvd and nvd.cisa_kev:
        action_bits.append("Flagged in CISA KEV")

    if perf:
        action_bits.append(f"IPS export performance rating: {perf}")
    if severity:
        action_bits.append(f"Severity context: {severity}")
    if desc:
        action_bits.append(desc)

    if sig_available == "Y":
        if overall_mode == "Prevent":
            rec = (
                "IPS signature is present and active in Prevent. "
                "Prioritize patching of the affected system. "
                "Validate gateway coverage and continue monitoring for effectiveness and false positives."
            )
        elif overall_mode == "Detect":
            rec = (
                "IPS signature is present and active in Detect. "
                "Prioritize patching of the affected system. "
                "Review events and traffic relevance, then consider promotion to Prevent if validation supports blocking."
            )
        elif overall_mode == "Inactive":
            rec = (
                "IPS signature exists but is inactive in the reviewed profile or gateway columns. "
                "Prioritize patching of the affected system. "
                "If traffic relevance is confirmed, consider enabling the protection in Detect mode before potential promotion to Prevent."
            )
        else:
            rec = (
                "IPS signature exists, but no applicable active mode was identified in the reviewed profile columns. "
                "Prioritize patching of the affected system. "
                "Validate protection applicability, profile assignment, and traffic relevance before considering Detect-mode enablement."
            )
    else:
        if network == "Y":
            rec = (
                "No IPS signature was identified in the provided profile export. "
                "Prioritize patching of the affected system. "
                "If a relevant IPS protection becomes available, consider enabling it in Detect mode for validation before potential promotion to Prevent."
            )
        else:
            rec = (
                "No IPS signature was identified in the provided profile export. "
                "Prioritize patching of the affected system. "
                "Firewall IPS mitigation is likely not applicable based on the lack of a network attack vector."
            )

    priority = classify_priority(nvd, sig_available, overall_mode)
    return ". ".join(action_bits).strip(), rec, priority


def enrich_rows(
    flat_df: pd.DataFrame,
    ips_lookup: Dict[str, List[dict]],
    profile_columns: List[str],
    use_nvd: bool,
    delay: float,
    cache_path: Optional[str],
) -> pd.DataFrame:
    unique_cves = sorted(flat_df[COL_CVES].dropna().astype(str).str.upper().unique())
    json_cache = load_cache(cache_path)
    nvd_cache: Dict[str, Optional[CveResult]] = {cve: None for cve in unique_cves}

    for cve in unique_cves:
        cached = json_cache.get(cve)
        if isinstance(cached, dict):
            nvd_cache[cve] = CveResult(**cached)

    if use_nvd:
        session = requests.Session()
        for cve in unique_cves:
            if nvd_cache.get(cve) is not None:
                continue
            try:
                result = fetch_nvd_cve(session, cve, delay)
                nvd_cache[cve] = result
                json_cache[cve] = result.__dict__
                save_cache(cache_path, json_cache)
                print(f"Fetched {cve}", file=sys.stderr)
            except Exception as exc:
                print(f"WARN: failed to fetch {cve}: {exc}", file=sys.stderr)

    save_cache(cache_path, json_cache)

    out_rows = []
    for _, row in flat_df.iterrows():
        row = row.copy()
        cve = clean_str(row.get(COL_CVES, "")).upper()
        nvd = nvd_cache.get(cve)
        matches = ips_lookup.get(cve, [])
        sig_available, overall_mode, coverage, perf, protections = summarize_signature_presence(matches, profile_columns)

        if nvd and nvd.base_score is not None:
            row[COL_SCORE] = nvd.base_score
        if nvd and nvd.network_exploitable:
            row[COL_NET] = nvd.network_exploitable
        if nvd and nvd.severity and not clean_str(row.get(COL_SEVERITY, "")):
            row[COL_SEVERITY] = nvd.severity

        row[COL_SIG] = sig_available
        row[COL_MODE] = overall_mode
        row[COL_FW] = coverage
        if perf:
            row[COL_PERF] = perf

        notes, rec, priority = choose_recommendation(row, nvd, sig_available, overall_mode, perf)
        row[COL_NOTES] = notes
        row[COL_REC] = rec
        row[COL_PRIORITY] = priority

        set_detect, set_prevent = derive_action_flags(
            sig_available=sig_available,
            overall_mode=overall_mode,
            network=clean_str(row.get(COL_NET, "")),
            priority=priority,
        )
        row[COL_DETECT_REVIEW] = set_detect
        row[COL_PREVENT_REVIEW] = set_prevent

        if set_prevent == "✓":
            action_type = "Promote / Enable Prevent"
        elif set_detect == "✓":
            action_type = "Enable / Review Detect"
        elif sig_available == "Y" and overall_mode == "Prevent":
            action_type = "Monitor Prevent Coverage"
        else:
            action_type = "Patch / Review"

        row[COL_ACTION_TYPE] = action_type
        row["NVD Description"] = nvd.description if nvd else ""
        row["NVD Attack Vector"] = nvd.attack_vector if nvd else ""
        row["NVD Severity"] = nvd.severity if nvd else ""
        row["NVD Weakness Summary"] = nvd.weakness_summary if nvd else ""
        row["CISA KEV"] = "Y" if (nvd and nvd.cisa_kev) else "N"
        row["Matching IPS Protections"] = protections
        row["Matching IPS Rows"] = len(matches)
        row["Manual Review Override"] = ""
        row["Manual Review Notes"] = ""
        out_rows.append(row)

    out_df = pd.DataFrame(out_rows)
    out_df[COL_MODE] = out_df[COL_MODE].replace("", "N/A").fillna("N/A")
    out_df.attrs.update(flat_df.attrs)
    return out_df


def build_grouped_sheet(flat_df: pd.DataFrame) -> pd.DataFrame:
    grouped_rows = []
    for iavm, group in flat_df.groupby(COL_IAVM, dropna=False):
        first = group.iloc[0].copy()
        first[COL_CVES] = "; ".join(group[COL_CVES].astype(str).tolist())
        first[COL_NET] = "; ".join(sorted(set(clean_str(v) for v in group[COL_NET] if clean_str(v))))
        first[COL_SIG] = "; ".join(sorted(set(clean_str(v) for v in group[COL_SIG] if clean_str(v))))
        first[COL_MODE] = "; ".join(sorted(set(clean_str(v) for v in group[COL_MODE] if clean_str(v))))
        first[COL_FW] = " | ".join(
            sorted(set(clean_str(v) for v in group[COL_FW] if clean_str(v) and clean_str(v) != "N/A"))
        ) or "N/A"
        first[COL_NOTES] = " | ".join(f"{r[COL_CVES]}: {clean_str(r[COL_NOTES])}" for _, r in group.iterrows())
        first[COL_REC] = " | ".join(f"{r[COL_CVES]}: {clean_str(r[COL_REC])}" for _, r in group.iterrows())
        first["CVEs In Block"] = len(group)
        first["Critical Priority CVEs"] = len(group[group[COL_PRIORITY] == "Critical"])
        first["High Priority CVEs"] = len(group[group[COL_PRIORITY] == "High"])
        first["Manual Review Needed"] = "Y" if any(group["Manual Review Override"].astype(str).str.strip()) else "N"
        grouped_rows.append(first)
    out = pd.DataFrame(grouped_rows)
    out.attrs.update(flat_df.attrs)
    return out


def build_original_row_summary(flat_df: pd.DataFrame) -> pd.DataFrame:
    summary_rows = []
    grouped = flat_df.groupby("Source Worksheet Row", dropna=False)

    for source_row, group in grouped:
        mode_values = set(group[COL_MODE].astype(str))
        if "Prevent" in mode_values:
            highest_mode = "Prevent"
        elif "Detect" in mode_values:
            highest_mode = "Detect"
        elif "Inactive" in mode_values:
            highest_mode = "Inactive"
        else:
            highest_mode = "N/A"

        priority_values = [clean_str(v) for v in group[COL_PRIORITY] if clean_str(v)]
        priority_order = ["Critical", "High", "Moderate", "Low"]
        ordered_priorities = [p for p in priority_order if p in set(priority_values)]

        recommendation_summary = " | ".join(
            f"{r[COL_CVES]}: {clean_str(r[COL_REC])}" for _, r in group.iterrows()
        )

        set_detect = "✓" if any(group[COL_DETECT_REVIEW].astype(str).eq("✓")) else ""
        set_prevent = "✓" if any(group[COL_PREVENT_REVIEW].astype(str).eq("✓")) else ""

        summary_rows.append(
            {
                "Source Worksheet Row": source_row,
                COL_IAVM: clean_str(group.iloc[0].get(COL_IAVM, "")),
                "Flattened CVEs": ", ".join(group[COL_CVES].astype(str).tolist()),
                "Any Signature Available": "Y" if any(group[COL_SIG].astype(str).eq("Y")) else "N",
                "Highest Active IPS Mode": highest_mode,
                "Coverage Summary": " | ".join(
                    sorted(
                        set(
                            clean_str(v)
                            for v in group[COL_FW]
                            if clean_str(v) and clean_str(v) != "N/A"
                        )
                    )
                ) or "N/A",
                "Priority Summary": ", ".join(ordered_priorities),
                "Recommendation Summary": recommendation_summary,
                COL_DETECT_REVIEW: set_detect,
                COL_PREVENT_REVIEW: set_prevent,
            }
        )

    out = pd.DataFrame(summary_rows)
    out.attrs.update(flat_df.attrs)
    return out


def build_findings_summary(flat_df: pd.DataFrame, grouped_df: pd.DataFrame, profile_columns: List[str]) -> pd.DataFrame:
    total_cves = len(flat_df)
    total_iavms = flat_df[COL_IAVM].nunique(dropna=True) if COL_IAVM in flat_df.columns else 0
    profile_text = ", ".join(profile_columns) if profile_columns else "Auto-detected / unspecified"

    priority_counts = flat_df[COL_PRIORITY].astype(str).value_counts() if COL_PRIORITY in flat_df.columns else pd.Series(dtype=int)
    mode_counts = flat_df[COL_MODE].astype(str).value_counts() if COL_MODE in flat_df.columns else pd.Series(dtype=int)

    recommended_detect = int((flat_df[COL_DETECT_REVIEW].astype(str) == "✓").sum()) if COL_DETECT_REVIEW in flat_df.columns else 0
    recommended_prevent = int((flat_df[COL_PREVENT_REVIEW].astype(str) == "✓").sum()) if COL_PREVENT_REVIEW in flat_df.columns else 0

    total_sig = int((flat_df[COL_SIG].astype(str) == "Y").sum()) if COL_SIG in flat_df.columns else 0
    total_no_sig = int((flat_df[COL_SIG].astype(str) == "N").sum()) if COL_SIG in flat_df.columns else 0
    total_network = int((flat_df[COL_NET].astype(str) == "Y").sum()) if COL_NET in flat_df.columns else 0
    total_kev = int((flat_df["CISA KEV"].astype(str) == "Y").sum()) if "CISA KEV" in flat_df.columns else 0

    critical_or_high = flat_df[flat_df[COL_PRIORITY].astype(str).isin(["Critical", "High"])] if COL_PRIORITY in flat_df.columns else flat_df.iloc[0:0]
    net_no_effective = flat_df[
        (flat_df[COL_NET].astype(str) == "Y")
        & (flat_df[COL_MODE].astype(str).isin(["Detect", "Inactive", "N/A", ""]))
    ] if COL_MODE in flat_df.columns and COL_NET in flat_df.columns else flat_df.iloc[0:0]
    inactive_matches = flat_df[
        (flat_df[COL_SIG].astype(str) == "Y")
        & (flat_df[COL_MODE].astype(str) == "Inactive")
    ] if COL_SIG in flat_df.columns and COL_MODE in flat_df.columns else flat_df.iloc[0:0]
    detect_matches = flat_df[
        (flat_df[COL_SIG].astype(str) == "Y")
        & (flat_df[COL_MODE].astype(str) == "Detect")
    ] if COL_SIG in flat_df.columns and COL_MODE in flat_df.columns else flat_df.iloc[0:0]

    def join_top_cves(df: pd.DataFrame, limit: int = 8) -> str:
        if df.empty or COL_CVES not in df.columns:
            return "None"
        vals = []
        for val in df[COL_CVES].astype(str):
            s = clean_str(val)
            if s and s not in vals:
                vals.append(s)
            if len(vals) >= limit:
                break
        return "; ".join(vals) if vals else "None"

    findings_narrative = [
        f"FAP reviewed {total_cves} CVEs across {total_iavms} IAVM notice blocks against the provided Check Point IPS export for profile(s): {profile_text}.",
        f"The review identified {priority_counts.get('Critical', 0)} Critical, {priority_counts.get('High', 0)} High, {priority_counts.get('Moderate', 0)} Moderate, and {priority_counts.get('Low', 0)} Low priority CVEs.",
        f"IPS protections were matched for {total_sig} CVEs, while {total_no_sig} CVEs had no matching protection in the provided export.",
        f"Of the reviewed CVEs, {total_network} were identified as network exploitable and {total_kev} were flagged in CISA KEV.",
    ]

    if recommended_detect or recommended_prevent:
        findings_narrative.append(
            f"Recommended IPS actions identified {recommended_detect} CVEs for Detect-first review and {recommended_prevent} CVEs for possible promotion to Prevent."
        )

    if not critical_or_high.empty:
        findings_narrative.append(
            f"Priority findings were driven primarily by CVEs lacking effective preventive IPS coverage. Example CVEs: {join_top_cves(critical_or_high)}."
        )
    if not inactive_matches.empty:
        findings_narrative.append(
            f"Some signatures exist in the export but are inactive in the reviewed profile(s). Example CVEs: {join_top_cves(inactive_matches)}."
        )
    if not detect_matches.empty:
        findings_narrative.append(
            f"Some matched protections are configured in Detect rather than Prevent and may warrant validation before promoting to blocking mode. Example CVEs: {join_top_cves(detect_matches)}."
        )

    recommended_actions = [
        "Prioritize patching or compensating controls for Critical and High CVEs, especially where no matching IPS protection was identified.",
        "Validate that the reviewed gateway/profile columns represent the relevant traffic path before treating IPS coverage as sufficient.",
        "Review signatures in Detect or Inactive mode and coordinate with firewall operations to determine whether they should be enabled, tuned, or promoted to Prevent.",
        "Coordinate with ICVM/CVM or system owners for follow-up remediation validation and residual risk tracking for CVEs lacking effective firewall-layer mitigation.",
    ]

    rows = []
    rows.append({"Section": "Assessment Metadata", "Metric": "Profiles Reviewed", "Value": profile_text, "Details": ""})
    rows.append({"Section": "Assessment Metadata", "Metric": "Total IAVM Notice Blocks", "Value": total_iavms, "Details": ""})
    rows.append({"Section": "Assessment Metadata", "Metric": "Total Flattened CVEs", "Value": total_cves, "Details": ""})

    for label in ["Critical", "High", "Moderate", "Low"]:
        rows.append({"Section": "Summary Metrics", "Metric": f"{label} Priority CVEs", "Value": int(priority_counts.get(label, 0)), "Details": ""})

    metric_rows = [
        ("Network Exploitable CVEs", total_network),
        ("CISA KEV CVEs", total_kev),
        ("CVEs With Matching IPS Protection", total_sig),
        ("CVEs Without Matching IPS Protection", total_no_sig),
        ("CVEs In Prevent", int(mode_counts.get("Prevent", 0))),
        ("CVEs Recommended to Set to Prevent", recommended_prevent),
        ("CVEs In Detect", int(mode_counts.get("Detect", 0))),
        ("CVEs Recommended to Set to Detect", recommended_detect),
        ("CVEs Inactive", int(mode_counts.get("Inactive", 0))),
        ("CVEs N/A", int(mode_counts.get("N/A", 0))),
    ]
    for metric, value in metric_rows:
        rows.append({"Section": "Summary Metrics", "Metric": metric, "Value": value, "Details": ""})

    for idx, text in enumerate(findings_narrative, start=1):
        rows.append({"Section": "FAP Assessment Summary", "Metric": f"Narrative {idx}", "Value": text, "Details": ""})

    for idx, text in enumerate(recommended_actions, start=1):
        rows.append({"Section": "Recommended Actions", "Metric": f"Action {idx}", "Value": text, "Details": ""})

    buckets = [
        ("No matching IPS protection in export", total_no_sig, join_top_cves(flat_df[flat_df[COL_SIG].astype(str) == "N"])),
        ("Network exploitable without Prevent coverage", len(net_no_effective), join_top_cves(net_no_effective)),
        ("Signatures present but Inactive", len(inactive_matches), join_top_cves(inactive_matches)),
        ("Signatures present in Detect", len(detect_matches), join_top_cves(detect_matches)),
        ("CISA KEV CVEs", total_kev, join_top_cves(flat_df[flat_df["CISA KEV"].astype(str) == "Y"])) if "CISA KEV" in flat_df.columns else ("CISA KEV CVEs", 0, "None"),
    ]
    for name, count, examples in buckets:
        rows.append({"Section": "Finding Buckets", "Metric": name, "Value": count, "Details": examples})

    return pd.DataFrame(rows)


def build_firewall_action_queue(flat_df: pd.DataFrame) -> pd.DataFrame:
    df = flat_df.copy()

    action_mask = (
        (df.get(COL_DETECT_REVIEW, "").astype(str) == "✓")
        | (df.get(COL_PREVENT_REVIEW, "").astype(str) == "✓")
        | (df.get(COL_PRIORITY, "").astype(str).isin(["Critical", "High"]))
    )

    queue = df[action_mask].copy()
    if queue.empty:
        cols = [
            "Source Worksheet Row",
            COL_IAVM,
            COL_CVES,
            COL_PRIORITY,
            COL_ACTION_TYPE,
            COL_SIG,
            COL_MODE,
            COL_FW,
            COL_REC,
        ]
        return pd.DataFrame(columns=cols)

    queue["Priority Sort"] = queue[COL_PRIORITY].map(
        {"Critical": 0, "High": 1, "Moderate": 2, "Low": 3}
    ).fillna(99)

    queue["Action Sort"] = queue[COL_ACTION_TYPE].map(
        {
            "Promote / Enable Prevent": 0,
            "Enable / Review Detect": 1,
            "Patch / Review": 2,
            "Monitor Prevent Coverage": 3,
        }
    ).fillna(99)

    queue = queue.sort_values(
        by=["Priority Sort", "Action Sort", "Source Worksheet Row", COL_CVES]
    )

    keep_cols = [
        "Source Worksheet Row",
        COL_IAVM,
        COL_CVES,
        COL_PRIORITY,
        COL_ACTION_TYPE,
        COL_SIG,
        COL_MODE,
        COL_FW,
        COL_DETECT_REVIEW,
        COL_PREVENT_REVIEW,
        COL_REC,
    ]
    return queue[keep_cols]


def set_preferred_widths(
    ws,
    header_row: int = 1,
    preferred_widths: Optional[Dict[str, int]] = None,
    default_min: int = 12,
    default_max: int = 40,
) -> None:
    preferred_widths = preferred_widths or {}
    for col_idx in range(1, ws.max_column + 1):
        header = clean_str(ws.cell(row=header_row, column=col_idx).value)
        letter = get_column_letter(col_idx)

        if header in preferred_widths:
            ws.column_dimensions[letter].width = preferred_widths[header]
            continue

        sample_values = [header]
        max_rows = min(ws.max_row, 300)
        for row_idx in range(2, max_rows + 1):
            sample_values.append(clean_str(ws.cell(row=row_idx, column=col_idx).value))

        width = min(max((len(v) for v in sample_values), default=default_min) + 2, default_max)
        ws.column_dimensions[letter].width = max(width, default_min)


def optimize_original_sheet_layout(ws, header_row: int = 1) -> None:
    preferred_widths = {
        COL_IAVM: 20,
        COL_DESC: 42,
        COL_SEVERITY: 14,
        COL_SCORE: 12,
        COL_FINDINGS: 22,
        COL_CVES: 26,
        COL_NET: 16,
        COL_PERF: 18,
        COL_SIG: 16,
        COL_MODE: 26,
        COL_FW: 28,
        COL_STATUS: 20,
        COL_OWNER: 16,
        COL_NOTES: 38,
        COL_REC: 42,
        "Any Signature Available": 18,
        "Highest Active IPS Mode": 20,
        "Coverage Summary": 28,
        "Priority Summary": 18,
        "Recommendation Summary": 44,
        COL_DETECT_REVIEW: 22,
        COL_PREVENT_REVIEW: 22,
    }
    set_preferred_widths(ws, header_row=header_row, preferred_widths=preferred_widths, default_min=12, default_max=44)


def append_df_to_ws(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    priority_col_idx = None
    sig_col_idx = None
    full_mode_col_idx = None

    for idx, col in enumerate(df.columns, start=1):
        if col == COL_PRIORITY:
            priority_col_idx = idx
        elif col == COL_SIG:
            sig_col_idx = idx
        elif col == COL_MODE:
            full_mode_col_idx = idx

    for row_vals in df.fillna("").itertuples(index=False, name=None):
        ws.append(list(row_vals))
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        if priority_col_idx:
            p = clean_str(ws.cell(row=row_num, column=priority_col_idx).value)
            if p == "Critical":
                ws.cell(row=row_num, column=priority_col_idx).fill = _CRIT_FILL
                ws.cell(row=row_num, column=priority_col_idx).font = Font(color="FFFFFF", bold=True)
            elif p == "High":
                ws.cell(row=row_num, column=priority_col_idx).fill = _BAD_FILL
            elif p == "Moderate":
                ws.cell(row=row_num, column=priority_col_idx).fill = _WARN_FILL
            else:
                ws.cell(row=row_num, column=priority_col_idx).fill = _GOOD_FILL

        if sig_col_idx:
            val = clean_str(ws.cell(row=row_num, column=sig_col_idx).value)
            if val == "N":
                ws.cell(row=row_num, column=sig_col_idx).fill = _WARN_FILL

        if full_mode_col_idx:
            val = clean_str(ws.cell(row=row_num, column=full_mode_col_idx).value)
            if val == "Prevent":
                ws.cell(row=row_num, column=full_mode_col_idx).fill = _GOOD_FILL
            elif val == "Detect":
                ws.cell(row=row_num, column=full_mode_col_idx).fill = _WARN_FILL
            elif val in {"Inactive", "N/A"}:
                ws.cell(row=row_num, column=full_mode_col_idx).fill = _BAD_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    set_preferred_widths(
        ws,
        preferred_widths={
            COL_DESC: 42,
            COL_NOTES: 40,
            COL_REC: 44,
            "NVD Description": 44,
            "NVD Weakness Summary": 36,
            "Matching IPS Protections": 36,
            COL_FW: 28,
            COL_CVES: 24,
        },
        default_min=12,
        default_max=50,
    )


def detect_header_row_in_sheet(ws, max_scan_rows: int = 15) -> int:
    best_row = 1
    best_score = -1
    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
        score = score_header_row(values)
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


def apply_instruction_formatting(ws) -> None:
    ws.sheet_view.showGridLines = False
    for cell_ref in ["A1", "A3", "A8", "A11", "A23", "A31"]:
        ws[cell_ref].font = Font(bold=True, size=12 if cell_ref != "A1" else 15)
    ws["A1"].fill = _SECTION_FILL
    ws["A3"].fill = _PRIMARY_FILL
    ws["A8"].fill = _SECTION_FILL
    ws["A11"].fill = _SECTION_FILL
    ws["A23"].fill = _PRIMARY_FILL
    ws["A31"].fill = _SECTION_FILL

    for cell in ws["A"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    ws.freeze_panes = "A2"


def write_output(
    original_cve_path: str,
    flat_df: pd.DataFrame,
    grouped_df: pd.DataFrame,
    row_summary_df: pd.DataFrame,
    findings_summary_df: pd.DataFrame,
    action_queue_df: pd.DataFrame,
    output_path: str,
) -> None:
    wb = load_workbook(original_cve_path)
    original_sheet = wb[wb.sheetnames[0]]
    original_sheet.title = "Original_CVE_Block"

    header_row = detect_header_row_in_sheet(original_sheet)
    data_start_row = header_row + 1

    summary_map = {}
    for _, row in row_summary_df.iterrows():
        summary_map[int(row["Source Worksheet Row"])] = row

    original_headers = [original_sheet.cell(row=header_row, column=col_idx).value for col_idx in range(1, original_sheet.max_column + 1)]
    new_cols = [
        "Any Signature Available",
        "Highest Active IPS Mode",
        "Coverage Summary",
        "Priority Summary",
        "Recommendation Summary",
        COL_DETECT_REVIEW,
        COL_PREVENT_REVIEW,
    ]

    start_col = len(original_headers) + 1
    for offset, col_name in enumerate(new_cols):
        cell = original_sheet.cell(row=header_row, column=start_col + offset)
        cell.value = col_name
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for excel_row in range(data_start_row, original_sheet.max_row + 1):
        summary = summary_map.get(excel_row)
        if summary is None:
            continue

        values = [
            clean_str(summary.get("Any Signature Available", "")),
            clean_str(summary.get("Highest Active IPS Mode", "")),
            clean_str(summary.get("Coverage Summary", "")),
            clean_str(summary.get("Priority Summary", "")),
            clean_str(summary.get("Recommendation Summary", "")),
            clean_str(summary.get(COL_DETECT_REVIEW, "")),
            clean_str(summary.get(COL_PREVENT_REVIEW, "")),
        ]

        for offset, value in enumerate(values):
            cell = original_sheet.cell(row=excel_row, column=start_col + offset)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    original_sheet.freeze_panes = f"A{data_start_row}"
    original_sheet.auto_filter.ref = original_sheet.dimensions
    optimize_original_sheet_layout(original_sheet, header_row=header_row)

    for ws_name in [
        "Instructions",
        "Firewall_Action_Queue",
        "Flat_CVE_View",
        "Grouped_IAVM_View",
        "Original_Row_Summary",
        "FAP_Findings_Summary",
    ]:
        if ws_name in wb.sheetnames:
            del wb[ws_name]

    ws_instr = wb.create_sheet("Instructions", 0)
    instructions = [
        ["FAP CVE IPS Correlation Workbook"],
        [""],
        ["PRIMARY USER ACTIONS:"],
        ["• Firewall Operators → Review 'Firewall_Action_Queue' first, then 'Original_Row_Summary' for row-mapped implementation context"],
        ["• Analysts → Use 'Flat_CVE_View' for detailed validation and supporting evidence"],
        ["• Leadership → Use 'FAP_Findings_Summary' for summary metrics and assessment narrative"],
        [""],
        ["Purpose:"],
        ["This workbook provides analysis of CVEs from IAVM notices against Check Point IPS protections to support firewall assurance, exposure analysis, and mitigation prioritization."],
        [""],
        ["Tabs Overview:"],
        [""],
        ["Instructions"],
        ["- Workbook guide and recommended tab order"],
        [""],
        ["Firewall_Action_Queue (PRIMARY OPS TAB)"],
        ["- Implementation-focused queue for firewall operators"],
        ["- Prioritizes Critical/High items and Detect/Prevent recommendations"],
        [""],
        ["Original_Row_Summary"],
        ["- Maps CVE analysis back to source worksheet rows"],
        ["- Best companion tab for row-by-row implementation traceability"],
        [""],
        ["FAP_Findings_Summary"],
        ["- Executive-level summary of findings"],
        ["- Metrics, risk breakdown, and narrative observations"],
        [""],
        ["Flat_CVE_View"],
        ["- One row per CVE"],
        ["- Detailed IPS coverage, priority, notes, and recommendations"],
        [""],
        ["Grouped_IAVM_View"],
        ["- Aggregated view by IAVM notice"],
        ["- Useful for IAVM-level rollups"],
        [""],
        ["Original_CVE_Block"],
        ["- Original workbook sheet with added summary columns"],
        [""],
        ["Notes:"],
        ["- IPS coverage reflects only the provided export/profile(s)"],
        ["- 'No signature found' refers to absence in the provided export, not necessarily the global protections database"],
        ["- Recommendations should be validated against operational context before implementation"],
    ]
    for row in instructions:
        ws_instr.append(row)
    apply_instruction_formatting(ws_instr)

    ws_action = wb.create_sheet("Firewall_Action_Queue")
    append_df_to_ws(ws_action, action_queue_df)
    set_preferred_widths(
        ws_action,
        preferred_widths={
            "Source Worksheet Row": 18,
            COL_IAVM: 18,
            COL_CVES: 18,
            COL_PRIORITY: 14,
            COL_ACTION_TYPE: 24,
            COL_SIG: 16,
            COL_MODE: 24,
            COL_FW: 28,
            COL_REC: 52,
            COL_DETECT_REVIEW: 22,
            COL_PREVENT_REVIEW: 22,
        },
        default_min=12,
        default_max=60,
    )

    ws_rows = wb.create_sheet("Original_Row_Summary")
    append_df_to_ws(ws_rows, row_summary_df)

    ws_findings = wb.create_sheet("FAP_Findings_Summary")
    append_df_to_ws(ws_findings, findings_summary_df)
    set_preferred_widths(
        ws_findings,
        preferred_widths={"Section": 24, "Metric": 34, "Value": 90, "Details": 60},
        default_min=14,
        default_max=90,
    )

    ws_flat = wb.create_sheet("Flat_CVE_View")
    append_df_to_ws(ws_flat, flat_df)

    ws_grouped = wb.create_sheet("Grouped_IAVM_View")
    append_df_to_ws(ws_grouped, grouped_df)

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Flatten IAVM CVE blocks, query NVD, correlate against Check Point IPS signatures, and preserve the original workbook."
    )
    parser.add_argument("--cve", required=True, help="Input CVE workbook (.xlsx)")
    parser.add_argument("--ips", required=True, help="Input IPS export (.csv or .xlsx)")
    parser.add_argument("--out", required=True, help="Output Excel workbook (.xlsx)")
    parser.add_argument(
        "--profile-columns",
        nargs="*",
        default=None,
        help="Optional IPS profile or gateway columns. If omitted, they are auto-detected.",
    )
    parser.add_argument("--skip-nvd", action="store_true", help="Skip NVD lookups and only do local IPS correlation.")
    parser.add_argument("--delay", type=float, default=1.2, help="Delay between NVD API requests in seconds. Default: 1.2")
    parser.add_argument("--cache", default="nvd_cache.json", help="JSON cache file for NVD results. Default: nvd_cache.json")
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    if os.path.splitext(args.cve)[1].lower() != ".xlsx":
        raise ValueError("--cve must be an .xlsx workbook for formatting preservation.")

    if os.path.splitext(args.out)[1].lower() != ".xlsx":
        raise ValueError("--out must end in .xlsx")

    df_cve = read_input_table(args.cve)
    df_ips = read_input_table(args.ips)

    flat_df = flatten_cve_blocks(df_cve)
    if COL_MODE not in flat_df.columns:
        flat_df[COL_MODE] = ""

    profile_columns = args.profile_columns or infer_profile_columns(df_ips)

    if not profile_columns:
        raise ValueError("No IPS profile columns were identified. Pass them explicitly with --profile-columns.")

    ips_lookup = build_ips_lookup(df_ips, profile_columns)
    enriched_flat = enrich_rows(
        flat_df,
        ips_lookup,
        profile_columns,
        use_nvd=not args.skip_nvd,
        delay=args.delay,
        cache_path=args.cache,
    )
    grouped_df = build_grouped_sheet(enriched_flat)
    row_summary_df = build_original_row_summary(enriched_flat)
    findings_summary_df = build_findings_summary(enriched_flat, grouped_df, profile_columns)
    action_queue_df = build_firewall_action_queue(enriched_flat)

    write_output(
        args.cve,
        enriched_flat,
        grouped_df,
        row_summary_df,
        findings_summary_df,
        action_queue_df,
        args.out,
    )

    print("Done")
    print(f"Output written to: {args.out}")
    print(f"Profile columns used: {', '.join(profile_columns)}")
    print(f"NVD cache file: {args.cache}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
