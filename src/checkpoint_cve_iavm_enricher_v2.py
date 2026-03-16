import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NVD_URL = "https://services.nvd.nist.gov/rest/json/cves/2.0"
CVE_PATTERN = re.compile(r"CVE-\d{4}-\d{4,7}", re.IGNORECASE)

COL_IAVM = "IAVM Notice"
COL_DESC = "Description"
COL_STIG = "STIG Severity"
COL_SCORE = "CVSS Base Score"
COL_FINDINGS = "Current Findings"
COL_CVES = "Related CVE(s)"
COL_NET = "Network Exploitable (Y/N)"
COL_PERF = "Performance Impact"
COL_SIG = "IPS Signature Available (Y/N)"
COL_MODE = "IPS Mode (Detect/Prevent/NA)"
COL_FW = "Firewall Coverage (Gateway/Cluster)"
COL_STATUS = "Mitigation Status"
COL_OWNER = "Owner"
COL_NOTES = "Notes / Actions"
COL_REC = "Recommendation"

CVE_BASE_COLUMNS = [
    COL_IAVM,
    COL_DESC,
    COL_STIG,
    COL_SCORE,
    COL_FINDINGS,
    COL_CVES,
    COL_NET,
    COL_PERF,
    COL_SIG,
    COL_MODE,
    COL_FW,
    COL_STATUS,
    COL_OWNER,
    COL_NOTES,
    COL_REC,
]

IPS_METADATA_COLUMNS = {
    "follow up",
    "protection",
    "industry release",
    "release date",
    "update date",
    "performance",
    "severity",
    "confidence",
    "optimized",
    "product",
    "sku",
    "package",
    "profile",
}

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
_WARN_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
_GOOD_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
_BAD_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")


@dataclass
class CveResult:
    cve_id: str
    description: str = ""
    base_score: Optional[float] = None
    severity: str = ""
    attack_vector: str = ""
    network_exploitable: str = ""
    weakness_summary: str = ""
    cisa_kev: bool = False



def clean_str(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()



def split_cves(value) -> List[str]:
    if pd.isna(value):
        return []
    return [c.upper() for c in CVE_PATTERN.findall(str(value))]



def read_input_table(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return pd.read_csv(path)
    return pd.read_excel(path)



def fill_down_iavm_blocks(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    context_cols = [c for c in [COL_IAVM, COL_DESC, COL_STIG, COL_SCORE, COL_FINDINGS, COL_STATUS, COL_OWNER] if c in out.columns]
    if context_cols:
        out[context_cols] = out[context_cols].ffill()
    return out



def flatten_cve_blocks(df: pd.DataFrame) -> pd.DataFrame:
    df = fill_down_iavm_blocks(df)
    rows = []
    for original_row_number, (_, row) in enumerate(df.iterrows(), start=2):
        cves = split_cves(row.get(COL_CVES, ""))
        if not cves:
            continue
        for block_cve_index, cve in enumerate(cves, start=1):
            r = row.to_dict()
            r[COL_CVES] = cve
            r["Source Worksheet Row"] = original_row_number
            r["CVE Position In Source Cell"] = block_cve_index
            rows.append(r)
    if not rows:
        raise ValueError("No CVE IDs were found in the 'Related CVE(s)' column.")
    out = pd.DataFrame(rows)
    for col in CVE_BASE_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    return out



def infer_profile_columns(df_ips: pd.DataFrame) -> List[str]:
    profile_cols = []
    for col in df_ips.columns:
        if clean_str(col).lower() not in IPS_METADATA_COLUMNS:
            profile_cols.append(col)
    return profile_cols



def normalize_mode(value: str) -> str:
    s = clean_str(value).lower()
    if not s:
        return ""
    if "prevent" in s:
        return "Prevent"
    if "detect" in s:
        return "Detect"
    if s in {"inactive", "n/a", "na", "disabled", "off", "false"}:
        return "Inactive"
    return clean_str(value)



def build_ips_lookup(df_ips: pd.DataFrame, profile_columns: List[str]) -> Dict[str, List[dict]]:
    if "Industry Release" not in df_ips.columns:
        raise ValueError("IPS CSV/XLSX must contain an 'Industry Release' column.")

    records = []
    for _, row in df_ips.iterrows():
        cves = split_cves(row.get("Industry Release", ""))
        if not cves:
            continue
        profile_modes = {col: normalize_mode(row.get(col, "")) for col in profile_columns}
        for cve in cves:
            records.append(
                {
                    "cve_id": cve,
                    "protection": clean_str(row.get("Protection", "")),
                    "industry_release": clean_str(row.get("Industry Release", "")),
                    "performance": clean_str(row.get("Performance", "")),
                    "severity": clean_str(row.get("Severity", "")),
                    "confidence": clean_str(row.get("Confidence", "")),
                    "optimized": normalize_mode(row.get("Optimized", "")),
                    "profiles": profile_modes,
                }
            )

    lookup: Dict[str, List[dict]] = {}
    for rec in records:
        lookup.setdefault(rec["cve_id"], []).append(rec)
    return lookup



def get_headers() -> Dict[str, str]:
    headers = {"User-Agent": "checkpoint-cve-iavm-enricher-v2/1.0"}
    api_key = os.getenv("NVD_API_KEY")
    if api_key:
        headers["apiKey"] = api_key
    return headers



def load_cache(path: Optional[str]) -> Dict[str, dict]:
    if not path or not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    return payload if isinstance(payload, dict) else {}



def save_cache(path: Optional[str], cache: Dict[str, dict]) -> None:
    if not path:
        return
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2, sort_keys=True)



def fetch_nvd_cve(session: requests.Session, cve_id: str, delay: float) -> CveResult:
    response = session.get(NVD_URL, params={"cveId": cve_id}, headers=get_headers(), timeout=30)
    response.raise_for_status()
    payload = response.json()
    vulns = payload.get("vulnerabilities", [])
    if not vulns:
        return CveResult(cve_id=cve_id)

    cve = vulns[0].get("cve", {})
    description = ""
    for item in cve.get("descriptions", []):
        if item.get("lang") == "en":
            description = item.get("value", "")
            break

    weakness_summary = "; ".join(
        sorted(
            {
                d.get("value", "")
                for weakness in cve.get("weaknesses", [])
                for d in weakness.get("description", [])
                if d.get("lang") == "en" and d.get("value")
            }
        )
    )

    metrics = cve.get("metrics", {})
    base_score = None
    severity = ""
    attack_vector = ""

    for key in ("cvssMetricV40", "cvssMetricV31", "cvssMetricV30", "cvssMetricV2"):
        if key not in metrics or not metrics[key]:
            continue
        metric = metrics[key][0]
        cvss = metric.get("cvssData", {})
        base_score = cvss.get("baseScore")
        severity = metric.get("baseSeverity") or cvss.get("baseSeverity") or metric.get("severity") or ""
        attack_vector = cvss.get("attackVector") or cvss.get("accessVector") or ""
        break

    cisa_kev = bool(cve.get("cisaExploitAdd"))
    network = "Y" if clean_str(attack_vector).upper() == "NETWORK" else "N"
    time.sleep(delay)
    return CveResult(
        cve_id=cve_id,
        description=description,
        base_score=base_score,
        severity=severity,
        attack_vector=attack_vector,
        network_exploitable=network,
        weakness_summary=weakness_summary,
        cisa_kev=cisa_kev,
    )



def summarize_signature_presence(matches: List[dict], profile_columns: List[str]) -> Tuple[str, str, str, str, str]:
    if not matches:
        return "N", "Signature Not Installed", "N/A", "", ""

    active_gateways = []
    all_modes = []
    perf_values = []
    protection_names = []

    for m in matches:
        protection_names.append(m.get("protection", ""))
        if m.get("performance"):
            perf_values.append(m["performance"])
        for col in profile_columns:
            mode = normalize_mode(m.get("profiles", {}).get(col, ""))
            if not mode or mode == "Inactive":
                continue
            active_gateways.append(f"{col}:{mode}")
            all_modes.append(mode)

    if "Prevent" in all_modes:
        overall_mode = "Prevent"
    elif "Detect" in all_modes:
        overall_mode = "Detect"
    else:
        overall_mode = "Inactive"

    coverage = ", ".join(sorted(set(active_gateways))) if active_gateways else "N/A"
    perf = ", ".join(sorted(set(v for v in perf_values if v)))
    prot = "; ".join(sorted(set(p for p in protection_names if p)))
    return "Y", overall_mode, coverage, perf, prot



def classify_priority(nvd: Optional[CveResult], sig_available: str, mode: str) -> str:
    if nvd and nvd.cisa_kev and nvd.network_exploitable == "Y":
        return "High"
    if nvd and (nvd.severity or "").upper() in {"CRITICAL", "HIGH"} and nvd.network_exploitable == "Y" and sig_available == "N":
        return "High"
    if sig_available == "Y" and mode == "Inactive":
        return "Medium"
    return "Normal"



def choose_recommendation(row: pd.Series, nvd: Optional[CveResult], sig_available: str, overall_mode: str, perf: str) -> Tuple[str, str, str]:
    desc = clean_str((nvd.description if nvd else "") or row.get(COL_DESC, ""))
    network = clean_str((nvd.network_exploitable if nvd else "") or row.get(COL_NET, ""))
    severity = clean_str((nvd.severity if nvd else "") or row.get(COL_STIG, ""))
    action_bits = []

    if network == "Y":
        action_bits.append("Network-reachable issue based on CVSS/NVD attack vector")
    elif network == "N":
        action_bits.append("No network attack vector identified from CVSS/NVD")
    else:
        action_bits.append("Network exploitability requires analyst review")

    if nvd and nvd.cisa_kev:
        action_bits.append("Flagged in CISA KEV")

    if sig_available == "Y":
        if overall_mode == "Prevent":
            rec = "IPS signature is present and active in Prevent. Validate gateway coverage, monitor for false positives, and patch the affected software on the normal change path."
        elif overall_mode == "Detect":
            rec = "IPS signature is present and active in Detect. Review events and performance impact, then consider moving to Prevent if the traffic path and tuning support it."
        else:
            rec = "IPS signature exists but is not active on the reviewed profile or gateway columns. Enable in Detect first, verify traffic relevance, then consider Prevent."
    else:
        if network == "Y":
            rec = "No installed IPS signature was found in the export. Search Check Point protections for an applicable signature, enable it in Detect first if available, and patch the affected product."
        else:
            rec = "Patch the affected software and review host or application-level controls. Firewall IPS mitigation may not be applicable for this CVE."

    if perf:
        action_bits.append(f"IPS export performance rating: {perf}")
    if severity:
        action_bits.append(f"Severity context: {severity}")
    if desc:
        action_bits.append(desc)

    priority = classify_priority(nvd, sig_available, overall_mode)
    return ". ".join(action_bits).strip(), rec, priority



def enrich_rows(flat_df: pd.DataFrame, ips_lookup: Dict[str, List[dict]], profile_columns: List[str], use_nvd: bool, delay: float, cache_path: Optional[str]) -> pd.DataFrame:
    unique_cves = sorted(flat_df[COL_CVES].dropna().astype(str).str.upper().unique())
    json_cache = load_cache(cache_path)
    nvd_cache: Dict[str, Optional[CveResult]] = {cve: None for cve in unique_cves}

    for cve in unique_cves:
        cached = json_cache.get(cve)
        if isinstance(cached, dict):
            nvd_cache[cve] = CveResult(**cached)

    if use_nvd:
        session = requests.Session()
        for cve in unique_cves:
            if nvd_cache.get(cve) is not None:
                continue
            try:
                result = fetch_nvd_cve(session, cve, delay)
                nvd_cache[cve] = result
                json_cache[cve] = result.__dict__
                print(f"Fetched {cve}", file=sys.stderr)
            except Exception as exc:
                print(f"WARN: failed to fetch {cve}: {exc}", file=sys.stderr)

    save_cache(cache_path, json_cache)

    out_rows = []
    for _, row in flat_df.iterrows():
        row = row.copy()
        cve = clean_str(row.get(COL_CVES, "")).upper()
        nvd = nvd_cache.get(cve)
        matches = ips_lookup.get(cve, [])
        sig_available, overall_mode, coverage, perf, protections = summarize_signature_presence(matches, profile_columns)

        if nvd and nvd.base_score is not None:
            row[COL_SCORE] = nvd.base_score
        if nvd and nvd.network_exploitable:
            row[COL_NET] = nvd.network_exploitable
        row[COL_SIG] = sig_available
        row[COL_MODE] = overall_mode if overall_mode else "N/A"
        row[COL_FW] = coverage
        if perf:
            row[COL_PERF] = perf
        notes, rec, priority = choose_recommendation(row, nvd, sig_available, overall_mode, perf)
        row[COL_NOTES] = notes
        row[COL_REC] = rec

        row["Analyst Priority"] = priority
        row["NVD Description"] = nvd.description if nvd else ""
        row["NVD Attack Vector"] = nvd.attack_vector if nvd else ""
        row["NVD Severity"] = nvd.severity if nvd else ""
        row["NVD Weakness Summary"] = nvd.weakness_summary if nvd else ""
        row["CISA KEV"] = "Y" if (nvd and nvd.cisa_kev) else "N"
        row["Matching IPS Protections"] = protections
        row["Matching IPS Rows"] = len(matches)
        row["Manual Review Override"] = ""
        row["Manual Review Notes"] = ""
        out_rows.append(row)

    return pd.DataFrame(out_rows)



def build_grouped_sheet(flat_df: pd.DataFrame) -> pd.DataFrame:
    grouped_rows = []
    for iavm, group in flat_df.groupby(COL_IAVM, dropna=False):
        first = group.iloc[0].copy()
        first[COL_CVES] = "; ".join(group[COL_CVES].astype(str).tolist())
        first[COL_NET] = "; ".join(sorted(set(clean_str(v) for v in group[COL_NET] if clean_str(v))))
        first[COL_SIG] = "; ".join(sorted(set(clean_str(v) for v in group[COL_SIG] if clean_str(v))))
        first[COL_MODE] = "; ".join(sorted(set(clean_str(v) for v in group[COL_MODE] if clean_str(v))))
        first[COL_FW] = " | ".join(sorted(set(clean_str(v) for v in group[COL_FW] if clean_str(v) and clean_str(v) != "N/A"))) or "N/A"
        first[COL_NOTES] = " | ".join(f"{row[COL_CVES]}: {clean_str(row[COL_NOTES])}" for _, row in group.iterrows())
        first[COL_REC] = " | ".join(f"{row[COL_CVES]}: {clean_str(row[COL_REC])}" for _, row in group.iterrows())
        first["CVEs In Block"] = len(group)
        first["High Priority CVEs"] = len(group[group["Analyst Priority"] == "High"])
        first["Manual Review Needed"] = "Y" if any(group["Manual Review Override"].astype(str).str.strip()) else "N"
        grouped_rows.append(first)
    return pd.DataFrame(grouped_rows)



def build_original_row_summary(flat_df: pd.DataFrame) -> pd.DataFrame:
    summary_rows = []
    grouped = flat_df.groupby("Source Worksheet Row", dropna=False)
    for source_row, group in grouped:
        summary_rows.append(
            {
                "Source Worksheet Row": source_row,
                COL_IAVM: clean_str(group.iloc[0].get(COL_IAVM, "")),
                "Flattened CVEs": "; ".join(group[COL_CVES].astype(str).tolist()),
                "Any Signature Available": "Y" if any(group[COL_SIG].astype(str).eq("Y")) else "N",
                "Highest Active IPS Mode": "Prevent" if any(group[COL_MODE].astype(str).eq("Prevent")) else ("Detect" if any(group[COL_MODE].astype(str).eq("Detect")) else "Inactive"),
                "Coverage Summary": " | ".join(sorted(set(clean_str(v) for v in group[COL_FW] if clean_str(v) and clean_str(v) != "N/A"))) or "N/A",
                "Priority Summary": "; ".join(sorted(set(clean_str(v) for v in group["Analyst Priority"] if clean_str(v)))),
                "Recommendation Summary": " | ".join(f"{r[COL_CVES]}: {clean_str(r[COL_REC])}" for _, r in group.iterrows()),
            }
        )
    return pd.DataFrame(summary_rows)



def append_df_to_ws(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    priority_col_idx = None
    sig_col_idx = None
    mode_col_idx = None
    for idx, col in enumerate(df.columns, start=1):
        if col == "Analyst Priority":
            priority_col_idx = idx
        elif col == COL_SIG:
            sig_col_idx = idx
        elif col == COL_MODE:
            mode_col_idx = idx

    for row_vals in df.fillna("").itertuples(index=False, name=None):
        ws.append(list(row_vals))
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if priority_col_idx:
            p = clean_str(ws.cell(row=row_num, column=priority_col_idx).value)
            if p == "High":
                ws.cell(row=row_num, column=priority_col_idx).fill = _BAD_FILL
            elif p == "Medium":
                ws.cell(row=row_num, column=priority_col_idx).fill = _WARN_FILL
            else:
                ws.cell(row=row_num, column=priority_col_idx).fill = _GOOD_FILL
        if sig_col_idx:
            val = clean_str(ws.cell(row=row_num, column=sig_col_idx).value)
            if val == "N":
                ws.cell(row=row_num, column=sig_col_idx).fill = _WARN_FILL
        if mode_col_idx:
            val = clean_str(ws.cell(row=row_num, column=mode_col_idx).value)
            if val == "Prevent":
                ws.cell(row=row_num, column=mode_col_idx).fill = _GOOD_FILL
            elif val in {"Inactive", "Signature Not Installed"}:
                ws.cell(row=row_num, column=mode_col_idx).fill = _WARN_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col in enumerate(df.columns, start=1):
        sample_values = [clean_str(col)] + [clean_str(v) for v in df[col].head(1000).tolist()]
        width = min(max(len(v) for v in sample_values) + 2, 70)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 12)



def write_output(original_cve_path: str, flat_df: pd.DataFrame, grouped_df: pd.DataFrame, row_summary_df: pd.DataFrame, output_path: str) -> None:
    wb = load_workbook(original_cve_path)
    original_sheet = wb[wb.sheetnames[0]]
    original_sheet.title = "Original_CVE_Block"

    for ws_name in ["Flat_CVE_View", "Grouped_IAVM_View", "Original_Row_Summary"]:
        if ws_name in wb.sheetnames:
            del wb[ws_name]

    ws_flat = wb.create_sheet("Flat_CVE_View")
    append_df_to_ws(ws_flat, flat_df)

    ws_grouped = wb.create_sheet("Grouped_IAVM_View")
    append_df_to_ws(ws_grouped, grouped_df)

    ws_rows = wb.create_sheet("Original_Row_Summary")
    append_df_to_ws(ws_rows, row_summary_df)

    wb.save(output_path)



def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Flatten IAVM CVE blocks, query NVD, correlate against Check Point IPS signatures, and preserve the original workbook.")
    parser.add_argument("--cve", required=True, help="Input CVE workbook (.xlsx)")
    parser.add_argument("--ips", required=True, help="Input IPS export (.csv or .xlsx)")
    parser.add_argument("--out", required=True, help="Output Excel workbook (.xlsx)")
    parser.add_argument("--profile-columns", nargs="*", default=None, help="Optional IPS profile or gateway columns. If omitted, they are auto-detected.")
    parser.add_argument("--skip-nvd", action="store_true", help="Skip NVD lookups and only do local IPS correlation.")
    parser.add_argument("--delay", type=float, default=0.7, help="Delay between NVD API requests in seconds. Default: 0.7")
    parser.add_argument("--cache", default="nvd_cache.json", help="JSON cache file for NVD results. Default: nvd_cache.json")
    return parser.parse_args()



def main() -> int:
    args = parse_args()
    if os.path.splitext(args.cve)[1].lower() != ".xlsx":
        raise ValueError("--cve must be an .xlsx workbook for formatting preservation in v2.")

    df_cve = read_input_table(args.cve)
    df_ips = read_input_table(args.ips)

    flat_df = flatten_cve_blocks(df_cve)
    profile_columns = args.profile_columns or infer_profile_columns(df_ips)
    if not profile_columns:
        raise ValueError("No IPS profile columns were identified. Pass them explicitly with --profile-columns.")

    ips_lookup = build_ips_lookup(df_ips, profile_columns)
    enriched_flat = enrich_rows(flat_df, ips_lookup, profile_columns, use_nvd=not args.skip_nvd, delay=args.delay, cache_path=args.cache)
    grouped_df = build_grouped_sheet(enriched_flat)
    row_summary_df = build_original_row_summary(enriched_flat)
    write_output(args.cve, enriched_flat, grouped_df, row_summary_df, args.out)

    print("Done")
    print(f"Output written to: {args.out}")
    print(f"Profile columns used: {', '.join(profile_columns)}")
    print(f"NVD cache file: {args.cache}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
